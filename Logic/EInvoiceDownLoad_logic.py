from selenium.webdriver.common.by import By
import time
import os
import threading
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from selenium import webdriver
import openpyxl
from datetime import datetime
import binascii
#import company
from Company.hiloCompany import HiloCompany
from Company.viettelCopany import ViettelCompany
from Company.misaEInvoice import MisaEInvoice
from Company.minvoiceCompany import MinvoiceCompany
from Company.ehoaDonCompany import EHoaDonCompany



class EInvoiceDownloadLogic:
    def __init__(self):
        threading.Thread.__init__(self)
        self.arCompany = []
        self.Dir_parameter = None
        self.Dir_Web = None
        self.datalistP = []
        self.download_folder = ''
    
    def UploadDriver(self,url_dic, url_Gdrive):
        gauth = GoogleAuth()
        Gdrive = GoogleDrive(gauth)
        option = webdriver.ChromeOptions()
        option.add_argument("user-data-dir=C:\\Users\\Admin\\AppData\\Local\\Google\\Chrome\\User Data")
        option.add_argument("--disable-blink-features=AutomationControlled")
        option.add_argument("no-sandbox")
        option.add_argument("--disable-gpu")
        option.add_argument("--disable-dev-shm-usage")
        option.add_experimental_option("useAutomationExtension", False)
        option.add_experimental_option("excludeSwitches", ["enable-automation"])  
        driver = webdriver.Chrome("chromedriver.exe",options=option)
        driver.get("https://accounts.google.com/o/oauth2/auth?client_id=585591886429-2imbpinhkcgp9mosgmmbp23ortijkncf.apps.googleusercontent.com&redirect_uri=http%3A%2F%2Flocalhost%3A8080%2F&scope=https%3A%2F%2Fwww.googleapis.com%2Fauth%2Fdrive&access_type=offline&response_type=code")
        client_json_path = 'D:/videoYoutube/python/AutoEmail/client_secret.json'    
        GoogleAuth.DEFAULT_SETTINGS['client_config_file'] = client_json_path
        driver.find_elements(By.XPATH,"//*[@id='view_container']/div/div/div[2]/div/div[1]/div/form/span/section/div/div/div/div/ul/li[1]/div/div[1]/div/div[2]/div[2]")[0].click()
        time.sleep(3)
        driver.find_elements(By.XPATH,"//*[@id='yDmH0d']/div[1]/div[1]/a")[0].click()
        time.sleep(1)
        driver.find_elements(By.XPATH,"//*[@id='yDmH0d']/div[1]/div[2]/p[2]/a")[0].click()
        time.sleep(3)
        driver.find_elements(By.XPATH,"//*[@id='submit_approve_access']/div/button/span")[0].click()
        try:
            print('Upload File GDrive Start')
            for f in os.listdir(url_dic):
                filename_os = os.path.join(url_dic,f)
                print(' File Name : ', filename_os)
                gfile = Gdrive.CreateFile({'parents':[{'id': url_Gdrive}],"title":f})
                gfile.SetContentFile(filename_os)
                print(' File Upload ')
                gfile.Upload()
                print('Upload File GDrive End')
        except print(0):
            print('Upload File GDrive Error')
            
    def Company(self,index):
        #driver = uc.Chrome(version_main=99)
        #os.open('"C:\Program Files\Google\Chrome\Application\chrome.exe" --user-data-dir=Profile3 --remote-debugging-port=9222')
        options = webdriver.ChromeOptions()
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("no-sandbox")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-notifications")
        options.add_experimental_option("useAutomationExtension", False)
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        PathOutput = self.CreateFolderByTime("./Output/")
        match index:
            case 0:
                print("[Start Company] https://sso.easyinvoice.vn/")
            case 1:
                self.ActionCompanyEHoaDon(options)
            case 2:
                self.ActionCompanyHilo(options,PathOutput)
            case 3:
                self.ActionCompanyMeinvoice(options)
            case 4:
                self.ActionCompanyMinvoce(options)
            case 5:
                self.ActionCompanyViettel(options)
            case 6:
                print("[Start] UPload All File To GDrive")
                #upload Driver
                #1Y3vHITzwi0icz0I0ijPGA9MmJze6CoLa
                self.UploadDriver("D:/videoYoutube/python/AutoEmail/img/","1Y3vHITzwi0icz0I0ijPGA9MmJze6CoLa")
            case default:
                print("Robot does not have this company data !")
    
    def read_parameter_By_company_logic(self,namesheet):
        book = openpyxl.load_workbook('./Invoice_Parameter.xlsx')
        sheet_Parameter = book[namesheet]   
        m_row = sheet_Parameter.max_row
        dir_sheet_data = {}
        for i in range(2, m_row + 1):
            cell_obj_key = sheet_Parameter.cell(row = i, column = 1)
            cell_obj_value = sheet_Parameter.cell(row = i, column = 2)
            dir_sheet_data[cell_obj_key.value] = cell_obj_value.value
            
        return dir_sheet_data
    
    def read_sheet_web_logic(self,namesheet):
        book = openpyxl.load_workbook('./Invoice_Parameter.xlsx')
        sheet_Parameter = book[namesheet]   
        m_row = sheet_Parameter.max_row
        dir_sheet_data = {}
        for i in range(2, m_row + 1):
            cell_obj_ID = sheet_Parameter.cell(row = i, column = 1)
            cell_obj_Link = sheet_Parameter.cell(row = i, column = 2)
            cell_obj_Active = sheet_Parameter.cell(row = i, column = 3)
            dir_sheet_data[i] = [cell_obj_ID.value,cell_obj_Link.value,cell_obj_Active.value]
            
        return dir_sheet_data
    
    def CreateFolderByTime(self,path):
        try:
            now = datetime.now() # current date and time
            date_time_Folder = now.strftime("%m%d%Y")
            print("Name Folder : ",date_time_Folder)
            newpath = path + date_time_Folder
            if not os.path.exists(newpath):
                os.makedirs(newpath)
            return newpath
        except print(0):
            pass

    def getDataByID(self,item):
        for i in self.Dir_parameter["Web"]:
            if int(i['val'].ID) == item:
                return i['val']
    
    def VNcode_custom(self,data):
        text = binascii.hexlify(data.encode('cp1258', errors='backslashreplace'))
        return binascii.unhexlify(text).decode('unicode-escape')

    def getCountHD(self,soup):
        countHD = 0
        for item in soup.find('tbody').find_all('tr')[:]:
            tds = item.find_all('td')[:]
            if (len(tds) > 0):
                countHD += 1
                for p in tds:
                    td = p.find_all('p')[:]
                    if (len(td) > 0):
                        for vl_custom in td:
                            if (vl_custom is not None):
                                vl_customs = self.VNcode_custom(str(vl_custom))
                                self.datalistP.append(vl_customs)
        return countHD
        
    def countPageSize(self,number_pagesize, number_HD):
        songuyen = 0
        if (number_HD > number_pagesize):
            c = int(number_HD % number_pagesize) # so du
            songuyen = (number_HD - c) / number_pagesize
        else:
            songuyen = 1
        return int(songuyen)  
                
    def find_current_page_size(self,soupPSize):
        for item in soupPSize.find_all('option', {"selected" : "selected"}):
            return int(item.text)

    def ActionCompanyHilo(self,options,PathOutput):
        print("[Start Company] https://hddt78.hilo.com.vn/")
        path = PathOutput + "/Hilo_Copany/Old_File/"
        driver = webdriver.Chrome("chromedriver.exe")
        driver = webdriver.Chrome(options=options)
        OjbHilo = HiloCompany(driver,path)
        name_folder = self.get_name_foldel()
        download_folder_c = ''.join([self.download_folder,'\\',name_folder,'\\Hilo_Copany\\Old_File\\'])
        params = {'behavior' : 'allow', 'downloadPath': download_folder_c}
        driver.execute_cdp_cmd('Page.setDownloadBehavior', params)
        OjbHilo.Action_Main(name_folder,download_folder_c)
        driver.close()
    
    def ActionCompanyViettel(self,options):   
        print("[Start Company] https://vinvoice.viettel.vn")
        driver = webdriver.Chrome("chromedriver.exe")
        driver = webdriver.Chrome(options=options)
        name_folder = self.get_name_foldel()
        download_folder_c = ''.join([self.download_folder,'\\',name_folder,'\\Vinvoice_Viettel_Copany\\Old_File\\'])
        params = {'behavior' : 'allow', 'downloadPath': download_folder_c}
        driver.execute_cdp_cmd('Page.setDownloadBehavior', params)
        OjbViettelCompany = ViettelCompany(driver)
        OjbViettelCompany.Action_Main(name_folder,download_folder_c)
        print("viettelCompany Success...")
        
    def ActionCompanyMinvoce(self,options):   
        print("[Start Company] https://*.minvoice.com.vn/")
        driver = webdriver.Chrome("chromedriver.exe")
        options.add_experimental_option('prefs', {
        "plugins.always_open_pdf_externally": True
        })
        driver = webdriver.Chrome(options=options)
        name_folder = self.get_name_foldel()
        download_folder_c = ''.join([self.download_folder,'\\',name_folder,'\\Minvoce_Copany\\Old_File\\'])
        params = {'behavior' : 'allow', 'downloadPath': download_folder_c}
        driver.execute_cdp_cmd('Page.setDownloadBehavior', params)
        OjbMinvoceCompany = MinvoiceCompany(driver)
        OjbMinvoceCompany.Action_Main(name_folder,download_folder_c)
        print("https://*.minvoice.com.vn/ Success...")
    
    def ActionCompanyEHoaDon(self,options):   
        print("[Start Company] https://*.ehoadon.com.vn/")
        driver = webdriver.Chrome("chromedriver.exe")
        options.add_experimental_option('prefs', {
        "plugins.always_open_pdf_externally": True
        })
        driver = webdriver.Chrome(options=options)
        name_folder = self.get_name_foldel()
        download_folder_c = ''.join([self.download_folder,'\\',name_folder,'\\EHoaDon_Copany\\'])
        params = {'behavior' : 'allow', 'downloadPath': download_folder_c}
        driver.execute_cdp_cmd('Page.setDownloadBehavior', params)
        OjbEHoaDonCompany = EHoaDonCompany(driver)
        OjbEHoaDonCompany.Action_Main(name_folder)
        print("https://*.ehoadon.com.vn/ Success...")
       
    def get_name_foldel(self):
        now = datetime.now() # current date and time
        date_time_Folder = now.strftime("%m%d%Y")
        return date_time_Folder
    
    def ActionCompanyMeinvoice(self,options):
        print("[Start Company] https://app3.meinvoice.vn/")
        driver = webdriver.Chrome("chromedriver.exe")
        driver = webdriver.Chrome(options=options)
        name_folder = self.get_name_foldel()
        download_folder_c = ''.join([self.download_folder,'\\',name_folder,'\\Misa_Einvoice_Copany\\Old_File\\'])
        params = {'behavior' : 'allow', 'downloadPath': download_folder_c}
        driver.execute_cdp_cmd('Page.setDownloadBehavior', params)
        OjbMisaEInvoice = MisaEInvoice(driver)
        OjbMisaEInvoice.Action_Main(name_folder,download_folder_c)
        print("Misa EInvoice Success...")
     
    def start_EInvoice(self):
        self.Dir_parameter = self.read_parameter_By_company_logic('Parameter')
        self.Dir_Web = self.read_sheet_web_logic('Web')
        arr_active_company = []
        self.download_folder = self.Dir_parameter['local_output_path']
        print('download_folder',self.download_folder)   
        for i in self.Dir_Web:
            id = int(self.Dir_Web[i][0])
            link = self.Dir_Web[i][1]
            active = int(self.Dir_Web[i][2])
            self.arCompany.append(link)
            if active == 1:
                arr_active_company.append(id)
        
        print('arr_active_company',arr_active_company)      
        return arr_active_company
            
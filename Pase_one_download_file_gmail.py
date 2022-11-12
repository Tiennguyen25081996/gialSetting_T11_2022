import time
import openpyxl
from Logic.ScrapeData_Download_Gmail import ScrapDataWeb_Download_Gmail
#rpa.thinktodoco@gmail.com
#rpa.thinkanddozone@gmail.com
NameCompany = ''
host = ''
username = ''
password = ''
download_folder = ''
arr_web_open_by_link = []
link_output = ''
page_size_cus = 0
start_date = ''
end_date = ''
arr_content_HDDT_c = []
arr_title_HDDT = []

arr_content_link = []
dir_sheet_Gmail = {}


def read_parameter_gmail():
    book = openpyxl.load_workbook('./Invoice_Parameter.xlsx')
    sheet_Parameter = book["Gmail"]   
    m_row = sheet_Parameter.max_row
    for i in range(2, m_row + 1):
        cell_obj_key = sheet_Parameter.cell(row = i, column = 1)
        cell_obj_value = sheet_Parameter.cell(row = i, column = 2)
        dir_sheet_Gmail[cell_obj_key.value] = cell_obj_value.value 
     
if __name__ == '__main__':
    read_parameter_gmail()
    NameCompany = dir_sheet_Gmail['NameCompany']
    host = dir_sheet_Gmail['host']
    username = dir_sheet_Gmail['username']
    password = dir_sheet_Gmail['password']
    download_folder = dir_sheet_Gmail['download_folder']
    link_output = dir_sheet_Gmail['link_output']
    page_size_cus = int(dir_sheet_Gmail['page_size_cus'])
    start_date = dir_sheet_Gmail['start_date']
    end_date = dir_sheet_Gmail['end_date']
    arr_content_HDDT_c = dir_sheet_Gmail['arr_content_HDDT_c'].split((','))
    arr_title_HDDT = dir_sheet_Gmail['arr_title_HDDT'].split((','))
    NameGmail = dir_sheet_Gmail['NameGmail']
    arr_title_HDDT.insert(0, NameCompany)
    arr_web_open_by_link = dir_sheet_Gmail['arr_web_open_by_link'].split((','))
    ojb_ScrapDataWeb_Download_Gmail = ScrapDataWeb_Download_Gmail(NameCompany,host,username,password,download_folder)
    # create folder Files
    name_folder = ojb_ScrapDataWeb_Download_Gmail.get_name_foldel()
    name_folder_files = ''.join([link_output,'/',name_folder,'/',NameGmail,'/Files/'])
    PathOutput = ojb_ScrapDataWeb_Download_Gmail.CreateFolderByPath(name_folder_files)
    print('PathOutput :',PathOutput)
    name_file_log_err = ''.join([name_folder,'_Error'])
    name_file_log_success = ''.join([name_folder,'_Success'])
    download_folder_c = ''.join([download_folder,'\\',name_folder,'\\',NameGmail,'\\Files\\'])
    log_folder = ''.join([download_folder,'\\',name_folder,'\\',NameGmail,'\\'])
    print('download_folder_c : ',download_folder_c)
    driver = ojb_ScrapDataWeb_Download_Gmail.setUp_Driver(download_folder_c)
    driver.get('https://mail.google.com/mail/u/0/?tab=rm#inbox')
    time.sleep(8)
    ojb_ScrapDataWeb_Download_Gmail.log_In_Gmail(driver)
    driver.maximize_window()
    time.sleep(20)
    #check tong so email 
    tong_email =  ojb_ScrapDataWeb_Download_Gmail.count_all_email_page_1st(driver)
    mood_email_page_c = tong_email % page_size_cus
    loop_email_page_c = (tong_email - mood_email_page_c) / page_size_cus
    time.sleep(8)
    ojb_ScrapDataWeb_Download_Gmail.search_email_by_date(driver,start_date,end_date)
    time.sleep(8)
    ojb_ScrapDataWeb_Download_Gmail.click_all_next_page(driver,log_folder,name_file_log_success,int(loop_email_page_c))
    time.sleep(3)
    count_all_email_c = ojb_ScrapDataWeb_Download_Gmail.count_all_email(driver)
    print('count_all_email_c',count_all_email_c)
    #back ve page_0
    for item in range(0,int(loop_email_page_c)):
        time.sleep(2)
        driver.back()
    print('Done Back Page 0 ...')
    url_c = driver.current_url
    driver.get(url_c)
    time.sleep(5)
    array_loop_mood = ojb_ScrapDataWeb_Download_Gmail.get_loop_mod_page_email(count_all_email_c,page_size_cus)
    loop_all_page= array_loop_mood[0]
    ojb_ScrapDataWeb_Download_Gmail.get_id_tr_span_arr(driver)
    ojb_ScrapDataWeb_Download_Gmail.Write_Log(log_folder,name_file_log_err,'========================== Danh sách email không có thông tin hoặc không tải được ============================ \n')
    ojb_ScrapDataWeb_Download_Gmail.Auto_click_download_each_email(driver,arr_title_HDDT,arr_content_HDDT_c,arr_web_open_by_link,log_folder,name_file_log_err,name_file_log_success,loop_all_page)
    time.sleep(2)
    driver.close()
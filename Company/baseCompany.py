from abc import ABC, abstractmethod
import autoit
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import os
import zipfile
import regex as re
import openpyxl
from datetime import datetime

class BaseCompany(ABC):
    def Login(self,driver,urlLogin,elTaxcode,elUserName,elPassWord,xpathelTaxcode,xpathelUserName,xpathelPassWord,xpathbtnLogin):
        driver.get(urlLogin)
        time.sleep(5)
        driver.maximize_window()
        if elTaxcode != "":
            taxcode = driver.find_elements(By.XPATH,xpathelTaxcode)[0]
            taxcode.click()
            time.sleep(1)
            taxcode.send_keys(elTaxcode)
        time.sleep(1)
        username = driver.find_elements(By.XPATH,xpathelUserName)[0]
        username.click()
        time.sleep(1)
        username.send_keys(elUserName)
        time.sleep(1)
        password = driver.find_elements(By.XPATH,xpathelPassWord)[0]
        password.click()
        time.sleep(1)
        password.send_keys(elPassWord)
        time.sleep(1)
        login = driver.find_elements(By.XPATH,xpathbtnLogin)[0]
        login.click()
        time.sleep(10)
    @abstractmethod
    def DownloadFile():
        pass
    @abstractmethod
    def PrintFile():
        pass    
    def PrintFile(self,nameFile):
        autoit.send("^p",0)
        time.sleep(5)
        autoit.send("{ENTER}",0)
        time.sleep(5)
        autoit.send(nameFile,0)
        time.sleep(3)
        autoit.send("{ENTER}",0)
        time.sleep(5)
    def Screenshot(self,driver, nameImg):
        img = ''
        img = nameImg + '.png'
        driver.save_screenshot(img)
        time.sleep(5)
    def zipFile(oldUrl,url):
        #'C:\\Stories\\Fantasy\\archive.zip'
        fantasy_zip = zipfile.ZipFile(url, 'w')
 
        for folder, subfolders, files in os.walk(oldUrl):
        
            for file in files:
                #if file.endswith('.pdf'):
                fantasy_zip.write(os.path.join(folder, file), os.path.relpath(os.path.join(folder,file), oldUrl), compress_type = zipfile.ZIP_DEFLATED)
        
        fantasy_zip.close()
        
    def UnZipFile(self,url, newUrl):
        fantasy_zip = zipfile.ZipFile(url)
        fantasy_zip.extractall(newUrl)
        fantasy_zip.close()
        
    def moveFile(self,fromURL, toURL):
        os.replace(fromURL, toURL)
        
    def renameFile(self,old_name,new_name):
        if os.path.isfile(new_name):
            print("The file already exists")
        else:
            # Rename the file
            os.rename(old_name, new_name)
    
    def delete_all_file_folder(self,path):
        arr = os.listdir(path)
        if (len(arr) > 0):
            for item in arr:
                path_c = ''.join([path,item])
                self.removeFile(path_c)
          
    def removeFile(self,name):
        os.remove(name)
    def get_all_name_file(self,path):
        arr = os.listdir(path)
        if (len(arr) > 0):
            return arr[0]
        else:
            return 0
    def no_accent_vietnamese(self,s):
        s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', s)
        s = re.sub(r'[ÀÁẠẢÃĂẰẮẶẲẴÂẦẤẬẨẪ]', 'A', s)
        s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', s)
        s = re.sub(r'[ÈÉẸẺẼÊỀẾỆỂỄ]', 'E', s)
        s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', s)
        s = re.sub(r'[ÒÓỌỎÕÔỒỐỘỔỖƠỜỚỢỞỠ]', 'O', s)
        s = re.sub(r'[ìíịỉĩ]', 'i', s)
        s = re.sub(r'[ÌÍỊỈĨ]', 'I', s)
        s = re.sub(r'[ùúụủũưừứựửữ]', 'u', s)
        s = re.sub(r'[ƯỪỨỰỬỮÙÚỤỦŨ]', 'U', s)
        s = re.sub(r'[ỳýỵỷỹ]', 'y', s)
        s = re.sub(r'[ỲÝỴỶỸ]', 'Y', s)
        s = re.sub(r'[Đ]', 'D', s)
        s = re.sub(r'[đ]', 'd', s)
        return s  
    def Setting_ask_download_chrome(driver):
        driver.get('chrome://settings/')
        time.sleep(5)
        driver.current_window_handle
        search_input = driver.switch_to.active_element
        time.sleep(1)
        search_input.send_keys('Download')
        time.sleep(2)
        search_input.send_keys(Keys.TAB)
        time.sleep(1)
        search_input.send_keys(Keys.TAB)
        time.sleep(1)
        search_input.send_keys(Keys.TAB)
        time.sleep(1)
        search_input.send_keys(Keys.TAB)
        time.sleep(1)
        search_input.send_keys(Keys.TAB)
        time.sleep(1)
        search_input.send_keys(Keys.TAB)
        time.sleep(1)
        search_input.send_keys(Keys.TAB)
        time.sleep(1)
        search_input.send_keys(Keys.TAB)
        time.sleep(1)
        search_input.send_keys(Keys.TAB)
        time.sleep(1)
        search_input.send_keys(Keys.ENTER)
        time.sleep(1)
        driver.back()
        driver.get('https://mail.google.com/mail/u/0/?tab=rm#inbox')
    def read_parameter_By_company(self,namesheet):
        book = openpyxl.load_workbook('./Invoice_Parameter.xlsx')
        sheet_Parameter = book[namesheet]   
        m_row = sheet_Parameter.max_row
        dir_sheet_data = {}
        for i in range(2, m_row + 1):
            cell_obj_key = sheet_Parameter.cell(row = i, column = 1)
            cell_obj_value = sheet_Parameter.cell(row = i, column = 2)
            dir_sheet_data[cell_obj_key.value] = cell_obj_value.value
            
        return dir_sheet_data   
    def create_folder_By_company(self,nameCompany):
        now = datetime.now() # current date and time
        date_time_Folder = now.strftime("%m%d%Y")
        newpath = "./Output/" + date_time_Folder + nameCompany
        print("Create Name Folder : ",newpath)
        if not os.path.exists(newpath):
            os.makedirs(newpath)